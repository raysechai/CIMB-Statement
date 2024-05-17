# change the 'file_name'

import pdfplumber
from openpyxl import Workbook

# set file name
file_name = 'C:/Users/rayse/OneDrive/Documents/Code/CIMB/CIMB_Statement/2024/eStatement20240118-000026822'

# Open the PDF file
with pdfplumber.open('{}.pdf'.format(file_name)) as pdf_file:
    
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    
    # Write the column headers to the first row
    ws.append(['Year', 'Month', 'Day', 'Balance', 'Expense', 'Category', 'Remarks', 'Description', 'Deposits', 'Category', 'Remarks', 'Description'])
    
    # Iterate over each page in the PDF file
    for page_num in range(len(pdf_file.pages)): ## range(xxx) # replace with the page number without table

        # Get the table on the current page
        table = pdf_file.pages[page_num].extract_table()

        # Check if table extraction was successful
        if table is None:
            print(f"Table extraction failed for page {page_num + 1}")
            continue

        # Iterate over each row in the table
        for row in table:
            # Check if the row is empty
            if not any(row):
                continue

            # Extract the date components
            date_components = row[0].split("/")
            if len(date_components) < 3:
                continue
            
            year = date_components[2]
            month_num = int(date_components[1])
            month_names = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sept', 'Oct', 'Nov', 'Dec']
            month = month_names[month_num]
            day = date_components[0]
            
            # Extract the expense and balance values
            expense = row[3]
            balance = row[6]
            
            # Extract the description and deposits values
            description = row[1]
            deposits = row[4] if row[4] else ''

            # Replace line breaks with spaces in description
            description = description.replace('\n', ' ')
            
            # Categorize based on the description
            if 'hktm-principal' in description.lower() or 'tng' in description.lower():
                expense_category = 'Tng'
            elif 'banking/wallets' in description.lower() or 'shopee top up' in description.lower() or 'shopeepay' in description.lower():
                expense_category = 'ShopeePay'
            elif 'shopee' in description.lower():
                expense_category = 'Shopee'
            elif 'atm withdrawal' in description.lower():
                expense_category = 'Withdraw'
            elif 'atm/debit card fee' in description.lower():
                expense_category = 'Services/Fee'
            elif 'foodpanda' in description.lower():
                expense_category = 'Foodpanda'
            elif 'hotlink' in description.lower() or 'shopee mobile' in description.lower():
                expense_category = 'Hotlink'
            elif 'gpay' in description.lower():
                expense_category = 'Grab'
            elif 'aham asset' in description.lower():
                expense_category = 'Versa'
            elif not expense:
                expense_category = '-'
            else:
                expense_category = ''
                
            # Categorize based on the deposit description
            deposit_category = ''
            if 'salary' in description.lower():
                deposit_category = 'Salary'
            elif 'credit interest' in description.lower():
                deposit_category = 'Services/Fees'
            elif 'claim' in description.lower():
                deposit_category = 'Claims'
            
            # Write the modified row to the Excel worksheet
            ws.append([year, month, day, balance, expense, expense_category, '', '', deposits, deposit_category, '', ''])
            
            # Check if the Withdrawal column is empty
            if row[4]:
                ws.cell(row=ws.max_row, column=12).value = description
            else:
                ws.cell(row=ws.max_row, column=8).value = description
            
    # Save the Excel workbook
    wb.save('{}.xlsx'.format(file_name))