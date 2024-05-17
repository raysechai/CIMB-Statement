# change the 'file_path' & 'file_name' & 'output_path'

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

file_path = 'C:/Users/rayse/OneDrive/Documents/Code/CIMB/CIMB_Statement/2023/'
file_name = os.path.join(file_path, '2023.xlsx')
output_path = os.path.join(file_path, '2023_calc.xlsx')

df = pd.read_excel(file_name)

# Get unique years, months, and categories
years = df.iloc[:, 0].unique()
months = df.iloc[:, 1].unique()
expense_categories = df.iloc[:, 5].unique()
deposit_categories = df.iloc[:, 9].unique()

# Create a new workbook
wb = Workbook()
ws = wb.active

row_index = 1
ws.cell(row=row_index, column=14, value="Total spent")
row_index += 1
# Print total unique combinations of year
for year in years:
    subset = df[(df.iloc[:, 0] == year)]
    if not subset.empty:
        balance_sum = subset.iloc[:, 4].str.replace(',', '').astype(float).sum()
        ws.cell(row=row_index, column=16, value=f"{year}")
        ws.cell(row=row_index, column=17, value=balance_sum)
        row_index += 1

row_index += 1
ws.cell(row=row_index, column=14, value="Total spent")
row_index += 1
# Print total by unique combinations of category
for category in expense_categories:
    if category == '-':  # skip the '-' category
        continue
    subset = df[(df.iloc[:, 5] == category)]
    if not subset.empty:
        balance_sum = subset.iloc[:, 4].str.replace(',', '').astype(float).sum()
        ws.cell(row=row_index, column=14, value=f"{category}")
        ws.cell(row=row_index, column=17, value=balance_sum)
        row_index += 1

row_index += 1
ws.cell(row=row_index, column=14, value="Total spent")
row_index += 1
# Print total by unique combinations of year and month
for year in years:
    for month in months:
        subset = df[(df.iloc[:, 0] == year) & (df.iloc[:, 1] == month)]
        if not subset.empty:
            balance_sum = subset.iloc[:, 4].str.replace(',', '').astype(float).sum()
            ws.cell(row=row_index, column=15, value=f"{month}")
            ws.cell(row=row_index, column=16, value=f"{year}")
            ws.cell(row=row_index, column=17, value=balance_sum)
            row_index += 1

row_index += 2
ws.cell(row=row_index, column=14, value="Total spent")
row_index += 1
# Print total by unique combinations of year and category
for year in years:
    for category in expense_categories:
        if category == '-':  # skip the '-' category
            continue
        subset = df[(df.iloc[:, 0] == year) & (df.iloc[:, 5] == category)]
        if not subset.empty:
            balance_sum = subset.iloc[:, 4].str.replace(',', '').astype(float).sum()
            ws.cell(row=row_index, column=14, value=f"{category}")
            ws.cell(row=row_index, column=16, value=f"{year}")
            ws.cell(row=row_index, column=17, value=balance_sum)
            row_index += 1

row_index += 1
ws.cell(row=row_index, column=14, value="Category")
# Gather category & print total by unique combinations of year and category
for category in expense_categories:
    row_index += 1
    if category == '-':  # skip the '-' category
        continue
    for year in years:
        subset = df[(df.iloc[:, 0] == year) & (df.iloc[:, 5] == category)]
        if not subset.empty:
            balance_sum = subset.iloc[:, 4].str.replace(',', '').astype(float).sum()
            ws.cell(row=row_index, column=14, value=f"{category}")
            ws.cell(row=row_index, column=16, value=f"{year}")
            ws.cell(row=row_index, column=17, value=balance_sum)
            row_index += 1

row_index += 2
# Print total by unique combinations of year, month and category
ws.cell(row=row_index, column=14, value="Total spent")
row_index += 1
for year in years:
    for month in months:
        for category in expense_categories:
            if category == '-':  # skip the '-' category
                continue
            subset = df[(df.iloc[:, 0] == year) & (df.iloc[:, 1] == month) & (df.iloc[:, 5] == category)]
            if not subset.empty:
                balance_sum = subset.iloc[:, 4].str.replace(',', '').astype(float).sum()
                ws.cell(row=row_index, column=14, value=f"{category}")
                ws.cell(row=row_index, column=15, value=f"{month}")
                ws.cell(row=row_index, column=16, value=f"{year}")
                ws.cell(row=row_index, column=17, value=balance_sum)
                row_index += 1

row_index += 1
# Gather category & print unique combinations of year, month and balance sum
ws.cell(row=row_index, column=14, value="Category")
for category in expense_categories:
    if category == '-':  # skip the '-' category
        continue
    row_index += 1
    for year in years:
        for month in months:
            subset = df[(df.iloc[:, 0] == year) & (df.iloc[:, 1] == month) & (df.iloc[:, 5] == category)]
            if not subset.empty:
                balance_sum = subset.iloc[:, 4].str.replace(',', '').astype(float).sum()
                ws.cell(row=row_index, column=14, value=f"{category}")
                ws.cell(row=row_index, column=15, value=f"{month}")
                ws.cell(row=row_index, column=16, value=f"{year}")
                ws.cell(row=row_index, column=17, value=balance_sum)
                row_index += 1

# Calculate total deposits

row_index = 1
ws.cell(row=row_index, column=19, value="Total deposits")
row_index += 1
# Print unique combinations of year and deposits sum
for year in years:
    subset = df[(df.iloc[:, 0] == year)]
    if not subset.empty:
        balance_sum = subset.iloc[:, 8].str.replace(',', '').astype(float).sum()
        ws.cell(row=row_index, column=21, value=f"{year}")
        ws.cell(row=row_index, column=22, value=balance_sum)
        row_index += 1

row_index += 1
ws.cell(row=row_index, column=19, value="Total deposits")
row_index += 1
# Print unique combinations of category, and deposits sum
for category in deposit_categories:
    if category == '-':  # skip the '-' category
        continue
    subset = df[(df.iloc[:, 9] == category)]
    if not subset.empty:
        balance_sum = subset.iloc[:, 8].str.replace(',', '').astype(float).sum()
        ws.cell(row=row_index, column=19, value=f"{category}")
        ws.cell(row=row_index, column=22, value=balance_sum)
        row_index += 1

row_index += 1
ws.cell(row=row_index, column=19, value="Total deposits")
row_index += 1
# Print unique combinations of year, month, and deposits sum
for year in years:
    for month in months:
        subset = df[(df.iloc[:, 0] == year) & (df.iloc[:, 1] == month)]
        if not subset.empty:
            balance_sum = subset.iloc[:, 8].str.replace(',', '').astype(float).sum()
            ws.cell(row=row_index, column=20, value=f"{month}")
            ws.cell(row=row_index, column=21, value=f"{year}")
            ws.cell(row=row_index, column=22, value=balance_sum)
            row_index += 1

row_index += 2
ws.cell(row=row_index, column=19, value="Total deposits")
row_index += 1
# Print total by unique combinations of year and category
for year in years:
    for category in deposit_categories:
        if category == '-':  # skip the '-' category
            continue
        subset = df[(df.iloc[:, 0] == year) & (df.iloc[:, 9] == category)]
        if not subset.empty:
            balance_sum = subset.iloc[:, 8].str.replace(',', '').astype(float).sum()
            ws.cell(row=row_index, column=19, value=f"{category}")
            ws.cell(row=row_index, column=21, value=f"{year}")
            ws.cell(row=row_index, column=22, value=balance_sum)
            row_index += 1

row_index += 1
ws.cell(row=row_index, column=19, value="Category")
# Print unique combinations of year, category, and deposits sum
for category in deposit_categories:
    row_index += 1
    if category == '-':  # skip the '-' category
        continue
    for year in years:
        subset = df[(df.iloc[:, 0] == year) & (df.iloc[:, 9] == category)]
        if not subset.empty:
            balance_sum = subset.iloc[:, 8].str.replace(',', '').astype(float).sum()
            ws.cell(row=row_index, column=19, value=f"{category}")
            ws.cell(row=row_index, column=21, value=f"{year}")
            ws.cell(row=row_index, column=22, value=balance_sum)
            row_index += 1

row_index += 2
ws.cell(row=row_index, column=19, value="Total deposits")
row_index += 1
# Print total by unique combinations of year, month and category
for year in years:
    for month in months:
        for category in deposit_categories:
            if category == '-':  # skip the '-' category
                continue
            subset = df[(df.iloc[:, 0] == year) & (df.iloc[:, 1] == month) & (df.iloc[:, 9] == category)]
            if not subset.empty:
                balance_sum = subset.iloc[:, 8].str.replace(',', '').astype(float).sum()
                ws.cell(row=row_index, column=19, value=f"{category}")
                ws.cell(row=row_index, column=20, value=f"{month}")
                ws.cell(row=row_index, column=21, value=f"{year}")
                ws.cell(row=row_index, column=22, value=balance_sum)
                row_index += 1

row_index += 1
ws.cell(row=row_index, column=19, value="Category")
# Gather category & print total by unique combinations of year, month and category
for category in deposit_categories:
    row_index += 1
    if category == '-':  # skip the '-' category
        continue
    for year in years:
        for month in months:
            subset = df[(df.iloc[:, 0] == year) & (df.iloc[:, 1] == month) & (df.iloc[:, 9] == category)]
            if not subset.empty:
                balance_sum = subset.iloc[:, 8].str.replace(',', '').astype(float).sum()
                ws.cell(row=row_index, column=19, value=f"{category}")
                ws.cell(row=row_index, column=20, value=f"{month}")
                ws.cell(row=row_index, column=21, value=f"{year}")
                ws.cell(row=row_index, column=22, value=balance_sum)
                row_index += 1

# Save the workbook
wb.save(output_path)
print("Analysis saved to Excel file.")